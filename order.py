import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm
from math import sqrt
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tabulate import tabulate
import warnings
import os

warnings.filterwarnings('ignore')
n = int(input("\nenter sample size : "))
mean = int(input("enter mean : "))
deviation_percent = float(input("Enter % deviation : "))
replenishment=int(input("enter the size of replenishment : "))
std = int(mean * (deviation_percent / 100))
L=[int(input("enter Lead time 1 : "))]
L.append(int(input("enter Lead time 2 : ")))
alpha=[float(input("enter % service rate 1 : "))/100]
alpha.append(float(input("enter % service rate 2 : "))/100)
print("\n")
zalpha=[norm.ppf(alpha[0])]
zalpha.append(norm.ppf(alpha[1]))
sqrtL=[sqrt(L[0])]
sqrtL.append(sqrt(L[1]))

safety_stock=[]
reorder_point=[]
Demand=[]
number_of_replenishments=[]

initial_stock=[[0 for _ in range(0,n)] for _ in range(0,4)]
final_stock=[[0 for _ in range(0,n)] for _ in range(0,4)]
final_stock_after_replenishment=[[0 for _ in range(0,n)] for _ in range(0,4)]
day=[x for x in range(1,n+1)]
order_size=[[0 for _ in range(0,n)] for _ in range(0,4)]
maximum_inventory=[]
for i in range(0,2):
    for j in range(0,2):
        safety_stock.append(zalpha[j]*sqrtL[i]*std)
        reorder_point.append((sqrtL[i]*sqrtL[i]*mean)+safety_stock[-1])

def random_nums(n,mean,std):  
    i=0
    x=[]
    while i<int(n/2):
        temp=2*mean
        while temp>(mean+std) or temp<(mean-std):
            temp=np.random.normal(mean,std)
        x.append(int(temp))
        i=i+1        
        
    y=[]

    for i in range(0,int(n/2)):
        
        if x[i]<mean:
            y.append((mean-x[i])+mean)
            
        elif x[i]>mean:
            y.append(mean-(x[i]-mean))
        else:
            y.append(x[i])
        y.append(x[i])
    nums=np.array(y)
    np.random.shuffle(nums)
    
    return [int(x) for x in nums]

de=random_nums(int(n),int(mean),int(std))
stock_outs=[["" for _ in range(0,n)] for _ in range(0,4)]
number_of_stockouts=[]

for x in range (0,4):
    Demand.append(de)
    initial_stock[x][0]=int(replenishment)
    number=0
    pos=0
    mark =True
    stnu=0
    for i in range(0,n):
        if i==pos and i!=0:
            final_stock_after_replenishment[x][i-1]=int(final_stock_after_replenishment[x][i-1]+replenishment)
            initial_stock[x][i]=int(final_stock_after_replenishment[x][i-1])
            mark=True
        final_stock[x][i]=int(initial_stock[x][i]-Demand[x][i])
        final_stock_after_replenishment[x][i]=final_stock[x][i]
        if (final_stock_after_replenishment[x][i]<reorder_point[x] or final_stock_after_replenishment[x][i]==reorder_point[x]) and i!=n-1 and mark==True:
            if x<2 and i<(n-1-L[0]):
                pos=i+L[0]+1
                order_size[x][i]=int(replenishment)
                number=number+1
                initial_stock[x][i+1]=int(final_stock_after_replenishment[x][i])
                mark=False
            elif x>1 and i<(n-1-L[1]):
                pos=i+L[1]+1
                order_size[x][i]=int(replenishment)
                number=number+1
                initial_stock[x][i+1]=int(final_stock_after_replenishment[x][i])
                mark=False
            else:
                initial_stock[x][i+1]=int(final_stock_after_replenishment[x][i])
        elif i!=n-1:
            initial_stock[x][i+1]=int(final_stock_after_replenishment[x][i])

            
        if final_stock[x][i]<1:
            stock_outs[x][i]="Yes"
            stnu=stnu+1
        
    number_of_replenishments.append(number)
    number_of_stockouts.append(stnu)
    maximum_inventory.append(np.max(initial_stock[x]))
result=[[0 for _ in range(0,9)] for _ in range(0,4)]
for i in range(0,4):
    if i%2==0:
        result[i][0]="{:.3f}".format(alpha[0])
        result[i][1]="{:.3f}".format(zalpha[0])
    else:
        result[i][0]="{:.3f}".format(alpha[1])
        result[i][1]="{:.3f}".format(zalpha[1])   
    if i<2:
        result[i][2]=L[0]
        result[i][3]="{:.2f}".format(sqrtL[0])
    else:
        result[i][2]=L[1]
        result[i][3]="{:.2f}".format(sqrtL[1])
    result[i][4]=int(safety_stock[i])
    result[i][5]=int(reorder_point[i])
    result[i][6]=int(np.max(initial_stock[i]))
    result[i][7]=int(number_of_replenishments[i])
    result[i][8]=int(number_of_stockouts[i])

results=pd.DataFrame(result,columns=["α","Zα","L","√L","Safety Stock","Reorder Point","Max Inventory","Replenishments","Stock outs"])
cx=np.stack((day,initial_stock[0],np.array(Demand[0]),final_stock_after_replenishment[0],order_size[0]),axis=1)
c1=pd.DataFrame(np.stack((day,initial_stock[0],np.array(Demand[0]),final_stock[0],stock_outs[0],final_stock_after_replenishment[0],order_size[0]),axis=1),index=None,columns=["Day","Initial stock","Demand","Final Stock","Stock outs","Final Stock After Replenishment","Order size"])
c2=pd.DataFrame(np.stack((day,initial_stock[1],np.array(Demand[1]),final_stock[1],stock_outs[1],final_stock_after_replenishment[1],order_size[1]),axis=1),index=None,columns=["Day","Initial stock","Demand","Final Stock","Stock outs","Final Stock After Replenishment","Order size"])
c3=pd.DataFrame(np.stack((day,initial_stock[2],np.array(Demand[2]),final_stock[2],stock_outs[2],final_stock_after_replenishment[2],order_size[2]),axis=1),index=None,columns=["Day","Initial stock","Demand","Final Stock","Stock outs","Final Stock After Replenishment","Order size"])
c4=pd.DataFrame(np.stack((day,initial_stock[3],np.array(Demand[3]),final_stock[3],stock_outs[3],final_stock_after_replenishment[3],order_size[3]),axis=1),index=None,columns=["Day","Initial stock","Demand","Final Stock","Stock outs","Final Stock After Replenishment","Order size"])
filename = 'Order.xlsx'

if not os.path.exists(filename):
    workbook = Workbook()
    sheet = workbook.active
    workbook.save(filename)
writer = pd.ExcelWriter('Order.xlsx', engine='openpyxl')
workbook = writer.book
df=[results,pd.merge(c1, c2, on='Day').merge(c3, on='Day').merge(c4, on='Day')]
if "Order_Strategy" not in workbook.sheetnames:
    workbook.create_sheet("Order_Strategy")
if "Conditions" not in workbook.sheetnames:
    workbook.create_sheet("Conditions")    
sheets=["Order_Strategy","Conditions"]
for i in range(0,2):
    worksheet=workbook[sheets[i]]
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None 
    df[i].to_excel(writer, sheet_name=sheets[i], startrow=0, startcol=0, header=True, index=False)
    writer.save()

print("\n ......................................................Results.....................................................\n")
print(tabulate(results.values.tolist(), headers=results.columns, tablefmt='grid',stralign='left'))



print("\n\n\n.....................................................Condition 1 ......................................................\n")
print(tabulate(c1.values.tolist(), headers=c1.columns, tablefmt='grid',stralign='left'))
print("\n\n\n.....................................................Condition 2 ......................................................\n")
print(tabulate(c2.values.tolist(), headers=c2.columns, tablefmt='grid',stralign='left'))
print("\n\n\n.....................................................Condition 3 ......................................................\n")
print(tabulate(c3.values.tolist(), headers=c3.columns, tablefmt='grid',stralign='left'))
print("\n\n\n.....................................................Condition 4 ......................................................\n")
print(tabulate(c4.values.tolist(), headers=c4.columns, tablefmt='grid',stralign='left'))
plt.hist(Demand[0],bins=n)
plt.title("Demand Distribution")
plt.xlabel('Demand')
plt.ylabel('Frequency')
plt.savefig('distribution.png')
plt.show()
